import datetime
import xlwt
from django.http import FileResponse, HttpResponse
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from xlwt.compat import xrange


def export_licens(modeladmin, request, queryset):
    data = {}

    filename = '{0}_{1}.xls'.format('Licenses', datetime.datetime.now().strftime('%d.%m.%Y_%H.%M.%S'))
    title = "Лицензии"
    columns = [
        (u"№ п/п", 7),
        (u"ID", 5),
        (u"ПО", 25),
        (u"Серийный номер", 32),
        (u"Количество", 12),
        (u"Дата получения", 17),
        (u"Дата установки", 17),
        (u"Город", 17),
        (u"Места установки", 30),

    ]
    row_num = 0
    row = []
    rows = []
    for obj in queryset:
        row_num += 1
        places = ''
        for place in obj.place_of_install_address.all():
            places += (str(place)) + '; '

        try:
            date_of_receiving = datetime.date.strftime(obj.date_of_receiving, '%d.%m.%Y')
        except:
            date_of_receiving = '-'

        try:
            date_of_install = datetime.date.strftime(obj.date_of_install, '%d.%m.%Y')
        except:
            date_of_install = '-'

        row = [
            row_num, obj.pk, obj.program.__str__(), obj.serial_number, obj.amount, date_of_receiving,
            date_of_install, obj.get_city_of_install_display(), str(places),
        ]
        rows.append(row)

    data = {'filename': filename, 'title': title, 'columns': columns, 'rows': rows, 'row_num': row_num}
    print(data)
    response = get_xlsx(data)
    # wb.save(response)
    return response


def export_certs(modeladmin, request, queryset):
    data = {}

    filename = '{0}_{1}.xls'.format('Certificates', datetime.datetime.now().strftime('%d.%m.%Y_%H.%M.%S'))
    title = "Сертификаты"
    columns = [
        (u"№ п/п", 7),
        (u"ID", 5),
        (u"Город", 15),
        (u"ФИО", 40),
        (u"Должность", 40),
        (u"Серийный номер", 30),
        (u"От", 10),
        (u"До", 10),
        (u"УЦ", 15),
        (u"СНИЛС", 13),
        (u"ИНН", 13),
        (u"ОГРН", 10),
        (u"Метки", 50),

    ]
    row_num = 0
    row = []
    rows = []
    for obj in queryset:
        row_num += 1
        tags = ''
        for tag in obj.tags.all():
            tags += (str(tag.name)) + ', '

        row = [
            row_num, obj.pk, obj.city, obj.name, obj.job, obj.serial_number, datetime.date.strftime(obj.start_date,
                                                                                                    '%d.%m.%Y'),
            datetime.date.strftime(obj.end_date, '%d.%m.%Y'), obj.uc, obj.snils, obj.inn, obj.ogrn, str(tags),
        ]
        rows.append(row)

    data = {'filename': filename, 'title': title, 'columns': columns, 'rows': rows, 'row_num': row_num}
    print(data)
    response = get_xlsx(data)
    # wb.save(response)
    return response


def get_xlsx(data):
    import openpyxl
    filename = data['filename']
    title = data['title']
    columns = data['columns']
    rows = data['rows']

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title

    # row_num = data['row_num']

    # columns = [
    #     (u"№ п/п", 10),
    #     (u"ID", 5),
    #     (u"Город", 15),
    #     (u"ФИО", 40),
    #     (u"Должность", 40),
    #     (u"Серийный номер", 30),
    #     (u"От", 10),
    #     (u"До", 10),
    #     (u"УЦ", 15),
    #     (u"СНИЛС", 13),
    #     (u"ИНН", 13),
    #     (u"ОГРН", 10),
    #     (u"Метки", 50),
    #
    # ]

    row_num = 0

    medium = Side(border_style="medium", color="000000")

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.font = Font(size=12, bold=True)
        c.alignment = Alignment(horizontal='center')
        c.border = Border(top=medium, bottom=medium, left=medium, right=medium)
        # Устанавливаем ширину столбца
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    print('rows\n', rows)
    for row in rows:
        print(row)
        row_num += 1
        #
        # tags = ''
        # for tag in obj.tags.all():
        #     tags += (str(tag.name)) + ', '

        # row = [
        #     row_num, obj.pk, obj.city, obj.name, obj.job, obj.serial_number, datetime.date.strftime(obj.start_date,
        #                                                                                             '%d.%m.%Y'),
        #     datetime.date.strftime(obj.end_date, '%d.%m.%Y'), obj.uc, obj.snils, obj.inn, obj.ogrn, str(tags),
        # ]

        thin = Side(border_style="thin", color="000000")

        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            c.alignment = Alignment(horizontal='left')
            c.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            # c.style.alignment.wrap_text = True

    wb.save(response)
    return response
